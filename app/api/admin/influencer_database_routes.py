"""
Influencer Master Database - Admin Routes (Endpoints 1-13 + Analytics)
Superadmin-only CRUD, bulk operations, export, Excel import, and analytics tracking for the influencer CRM.
"""
import io
import logging
import time
from datetime import datetime, timezone
from typing import Optional, List
from uuid import UUID

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.responses import StreamingResponse
from sqlalchemy import text
from sqlalchemy.ext.asyncio import AsyncSession

from app.database.optimized_pools import get_db_optimized as get_db
from app.middleware.auth_middleware import require_admin
from app.models.influencer_database import (
    AddInfluencerRequest,
    BulkImportRequest,
    BulkPricingRequest,
    BulkTagRequest,
    ExportRequest,
    InfluencerUpdateRequest,
)
from app.services.influencer_database_service import InfluencerDatabaseService

router = APIRouter(tags=["Influencer Database"])
logger = logging.getLogger(__name__)


# =============================================================================
# 1. LIST INFLUENCERS
# =============================================================================

@router.get("/influencers/database")
async def list_influencers(
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    search: Optional[str] = Query(None),
    status_filter: Optional[str] = Query(None, alias="status"),
    tier: Optional[str] = Query(None),
    tags: Optional[str] = Query(None, description="Comma-separated tags"),
    categories: Optional[str] = Query(None, description="Comma-separated categories"),
    min_followers: Optional[int] = Query(None, ge=0),
    max_followers: Optional[int] = Query(None, ge=0),
    engagement_min: Optional[float] = Query(None, ge=0),
    engagement_max: Optional[float] = Query(None, ge=0),
    is_verified: Optional[bool] = Query(None),
    has_pricing: Optional[bool] = Query(None),
    sort_by: str = Query("created_at"),
    sort_order: str = Query("desc"),
    db: AsyncSession = Depends(get_db),
    current_user=Depends(require_admin()),
):
    """List all influencers with pagination and filters."""
    try:
        tag_list = [t.strip() for t in tags.split(",")] if tags else None
        category_list = [c.strip() for c in categories.split(",")] if categories else None
        result = await InfluencerDatabaseService.list_influencers(
            db,
            page=page,
            page_size=page_size,
            search=search,
            status=status_filter,
            tier=tier,
            tags=tag_list,
            categories=category_list,
            min_followers=min_followers,
            max_followers=max_followers,
            engagement_min=engagement_min,
            engagement_max=engagement_max,
            is_verified=is_verified,
            has_pricing=has_pricing,
            sort_by=sort_by,
            sort_order=sort_order,
        )
        return {"success": True, "data": result}
    except Exception as e:
        logger.error(f"Error listing influencers: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# =============================================================================
# 2. GET ALL TAGS
# =============================================================================

@router.get("/influencers/tags")
async def get_all_tags(
    db: AsyncSession = Depends(get_db),
    current_user=Depends(require_admin()),
):
    """Get all unique tags used across influencers."""
    try:
        tags = await InfluencerDatabaseService.get_all_tags(db)
        return {"success": True, "data": {"tags": tags}}
    except Exception as e:
        logger.error(f"Error fetching tags: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# =============================================================================
# 3. ADD SINGLE INFLUENCER
# =============================================================================

@router.post("/influencers/add")
async def add_influencer(
    request: AddInfluencerRequest,
    db: AsyncSession = Depends(get_db),
    current_user=Depends(require_admin()),
):
    """Add a single influencer by Instagram username."""
    try:
        record = await InfluencerDatabaseService.add_influencer(
            db,
            username=request.username,
            added_by=current_user.id,
            categories=request.categories,
            tags=request.tags,
            notes=request.notes,
            status=request.status.value if request.status else "active",
            tier=request.tier,
            cost_pricing=request.cost_pricing.dict(exclude_none=True) if request.cost_pricing else None,
            sell_pricing=request.sell_pricing.dict(exclude_none=True) if request.sell_pricing else None,
        )
        return {"success": True, "data": {"influencer": record}, "message": f"@{request.username} added"}
    except ValueError as e:
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail=str(e))
    except Exception as e:
        logger.error(f"Error adding influencer @{request.username}: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# =============================================================================
# 4. BULK IMPORT
# =============================================================================

@router.post("/influencers/bulk-import")
async def bulk_import(
    request: BulkImportRequest,
    db: AsyncSession = Depends(get_db),
    current_user=Depends(require_admin()),
):
    """Bulk import influencers by username list."""
    try:
        result = await InfluencerDatabaseService.bulk_import(
            db, request.usernames, current_user.id
        )
        return {"success": True, "data": result}
    except Exception as e:
        logger.error(f"Error bulk importing: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# =============================================================================
# 5. BULK TAG
# =============================================================================

@router.post("/influencers/bulk-tag")
async def bulk_tag(
    request: BulkTagRequest,
    db: AsyncSession = Depends(get_db),
    current_user=Depends(require_admin()),
):
    """Bulk add or remove tags on multiple influencers."""
    try:
        count = await InfluencerDatabaseService.bulk_tag(
            db,
            influencer_ids=request.influencer_ids,
            tags=request.tags,
            action=request.action.value,
        )
        return {"success": True, "data": {"updated_count": count}}
    except Exception as e:
        logger.error(f"Error bulk tagging: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# =============================================================================
# 6. BULK PRICING
# =============================================================================

@router.post("/influencers/bulk-pricing")
async def bulk_pricing(
    request: BulkPricingRequest,
    db: AsyncSession = Depends(get_db),
    current_user=Depends(require_admin()),
):
    """Bulk update pricing for multiple influencers."""
    try:
        updates = [u.dict(exclude_none=True) for u in request.updates]
        result = await InfluencerDatabaseService.bulk_pricing(db, updates)
        return {"success": True, "data": result}
    except Exception as e:
        logger.error(f"Error bulk pricing: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# =============================================================================
# 7. EXPORT
# =============================================================================

@router.post("/influencers/export")
async def export_influencers(
    request: ExportRequest,
    db: AsyncSession = Depends(get_db),
    current_user=Depends(require_admin()),
):
    """Export influencers to CSV or JSON."""
    try:
        content, media_type, filename = await InfluencerDatabaseService.export_influencers(
            db,
            format=request.format.value,
            fields=request.fields,
            scope=request.scope.value,
            selected_ids=request.selected_ids,
            filters=request.filters,
        )

        def iter_content():
            yield content

        return StreamingResponse(
            iter_content(),
            media_type=media_type,
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )
    except Exception as e:
        logger.error(f"Error exporting: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# =============================================================================
# 8. GET DETAILED (by UUID or username)
# =============================================================================

@router.get("/influencers/{identifier}/detailed")
async def get_influencer_detailed(
    identifier: str,
    db: AsyncSession = Depends(get_db),
    current_user=Depends(require_admin()),
):
    """Get detailed influencer data by UUID or username. DB-only lookup (influencer_database then profiles table)."""
    try:
        result = await InfluencerDatabaseService.get_detailed(db, identifier)
        return {"success": True, "data": result}
    except ValueError as e:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail=str(e))
    except Exception as e:
        logger.error(f"Error getting influencer {identifier}: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# =============================================================================
# 9. UPDATE INFLUENCER
# =============================================================================

@router.put("/influencers/{influencer_id}")
async def update_influencer(
    influencer_id: UUID,
    request: InfluencerUpdateRequest,
    db: AsyncSession = Depends(get_db),
    current_user=Depends(require_admin()),
):
    """Update an influencer's CRM data and pricing."""
    try:
        data = request.dict(exclude_none=True)
        # Convert nested pydantic to dict
        if "cost_pricing" in data and data["cost_pricing"]:
            data["cost_pricing"] = (
                request.cost_pricing.dict(exclude_none=True)
                if request.cost_pricing else {}
            )
        if "sell_pricing" in data and data["sell_pricing"]:
            data["sell_pricing"] = (
                request.sell_pricing.dict(exclude_none=True)
                if request.sell_pricing else {}
            )
        result = await InfluencerDatabaseService.update_influencer(db, influencer_id, data)
        return {"success": True, "data": {"influencer": result}}
    except ValueError as e:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail=str(e))
    except Exception as e:
        logger.error(f"Error updating influencer {influencer_id}: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# =============================================================================
# 10. DELETE INFLUENCER
# =============================================================================

@router.delete("/influencers/{influencer_id}")
async def delete_influencer(
    influencer_id: UUID,
    db: AsyncSession = Depends(get_db),
    current_user=Depends(require_admin()),
):
    """Delete an influencer from the database."""
    try:
        deleted = await InfluencerDatabaseService.delete_influencer(db, influencer_id)
        if not deleted:
            raise HTTPException(status_code=404, detail="Influencer not found")
        return {"success": True, "message": "Influencer deleted"}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error deleting influencer {influencer_id}: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# =============================================================================
# 11a. BATCH ANALYTICS STATUS (for real-time polling)
# =============================================================================

@router.get("/influencer-database/analytics-status")
async def get_analytics_status(
    ids: str = Query(..., description="Comma-separated IMD IDs"),
    db: AsyncSession = Depends(get_db),
    current_user=Depends(require_admin()),
):
    """Return analytics status for multiple influencers in one call (polling endpoint)."""
    try:
        id_list = [id_val.strip() for id_val in ids.split(",") if id_val.strip()]
        if not id_list:
            return {"success": True, "data": []}

        # Build parameterized query for the list of UUIDs
        placeholders = ", ".join(f":id_{i}" for i in range(len(id_list)))
        params = {f"id_{i}": id_val for i, id_val in enumerate(id_list)}

        result = await db.execute(
            text(f"""
                SELECT id, analytics_status, analytics_progress,
                       analytics_progress_message, analytics_error,
                       analytics_completed_at
                FROM influencer_database
                WHERE id IN ({placeholders})
            """),
            params,
        )
        rows = [dict(r) for r in result.mappings().fetchall()]

        # Convert UUID and datetime to strings for JSON serialization
        for row in rows:
            row["id"] = str(row["id"])
            if row.get("analytics_completed_at"):
                row["analytics_completed_at"] = row["analytics_completed_at"].isoformat()

        return {"success": True, "data": rows}
    except Exception as e:
        logger.error(f"Error fetching analytics status: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# =============================================================================
# 11b. TRIGGER / RETRY ANALYTICS
# =============================================================================

@router.post("/influencer-database/{influencer_id}/trigger-analytics")
async def trigger_analytics(
    influencer_id: UUID,
    db: AsyncSession = Depends(get_db),
    current_user=Depends(require_admin()),
):
    """Manually trigger or retry analytics for an influencer."""
    try:
        # Get current record
        result = await db.execute(
            text("SELECT id, username, analytics_status FROM influencer_database WHERE id = CAST(:id AS uuid)"),
            {"id": str(influencer_id)},
        )
        record = result.mappings().fetchone()
        if not record:
            raise HTTPException(status_code=404, detail="Influencer not found")

        # Guard: don't double-queue
        if record["analytics_status"] == "processing":
            raise HTTPException(status_code=409, detail="Analytics already in progress")
        if record["analytics_status"] == "queued":
            raise HTTPException(status_code=409, detail="Analytics already queued")

        # Queue new analytics job (with unique idempotency key for retry)
        from app.core.job_queue import job_queue, QueueType

        job = await job_queue.enqueue_job(
            user_id=str(current_user.id),
            job_type='imd_creator_analytics',
            params={'username': record["username"], 'influencer_db_id': str(record["id"])},
            queue_type=QueueType.API_QUEUE,
            user_tier='enterprise',  # Admin operation
            idempotency_key=f"imd_analytics_{record['username']}_{int(time.time())}",
        )

        if not job.get('success', True):
            raise HTTPException(status_code=503, detail=f"Job queue rejected: {job.get('message', 'unknown')}")

        now = datetime.now(timezone.utc)
        job_id = job.get('job_id') or job.get('id')
        await db.execute(
            text("""
                UPDATE influencer_database
                SET analytics_status = 'queued',
                    analytics_job_id = CAST(:job_id AS uuid),
                    analytics_queued_at = :queued_at,
                    analytics_error = NULL,
                    analytics_progress = 0,
                    analytics_progress_message = NULL,
                    updated_at = :queued_at
                WHERE id = CAST(:id AS uuid)
            """),
            {"id": str(record["id"]), "job_id": job_id, "queued_at": now}
        )
        await db.commit()

        updated = {"id": str(record["id"]), "status": "queued", "job_id": job_id}
        return {"success": True, "data": {"status": "queued", "influencer": updated}}

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error triggering analytics for {influencer_id}: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# =============================================================================
# 11. REFRESH ANALYTICS
# =============================================================================

@router.post("/influencers/{influencer_id}/refresh")
async def refresh_influencer(
    influencer_id: UUID,
    db: AsyncSession = Depends(get_db),
    current_user=Depends(require_admin()),
):
    """Re-fetch Instagram data for an influencer."""
    try:
        result = await InfluencerDatabaseService.refresh_influencer(db, influencer_id)
        return {"success": True, "data": {"influencer": result}, "message": "Analytics refreshed"}
    except ValueError as e:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail=str(e))
    except Exception as e:
        logger.error(f"Error refreshing influencer {influencer_id}: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# =============================================================================
# 12. EXCEL TEMPLATE DOWNLOAD
# =============================================================================

TEMPLATE_COLUMNS = [
    "username", "status", "tier", "categories", "tags", "internal_notes",
]

# Legacy 30-column format for backward compatibility
LEGACY_TEMPLATE_COLUMNS = [
    "username", "full_name", "biography", "is_verified",
    "followers_count", "following_count", "posts_count",
    "engagement_rate", "avg_likes", "avg_comments", "avg_views",
    "status", "tier", "categories", "tags", "internal_notes",
    "cost_post_usd", "cost_story_usd", "cost_reel_usd",
    "cost_carousel_usd", "cost_video_usd", "cost_bundle_usd", "cost_monthly_usd",
    "sell_post_usd", "sell_story_usd", "sell_reel_usd",
    "sell_carousel_usd", "sell_video_usd", "sell_bundle_usd", "sell_monthly_usd",
]


@router.get("/influencer-database/template/download")
async def download_excel_template(
    current_user=Depends(require_admin()),
):
    """Download .xlsx template for influencer bulk import."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = Workbook()
        ws = wb.active
        ws.title = "Influencers"

        # Header styling
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")

        for col_idx, col_name in enumerate(TEMPLATE_COLUMNS, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            ws.column_dimensions[cell.column_letter].width = max(len(col_name) + 4, 14)

        # Add example row
        example = [
            "johndoe", "active", "micro", "Fashion, Lifestyle",
            "dubai, travel", "Great engagement with UAE audience",
        ]
        for col_idx, val in enumerate(example, 1):
            ws.cell(row=2, column=col_idx, value=val)

        # Instructions sheet
        inst = wb.create_sheet("Instructions")
        instructions = [
            ["Column", "Description", "Required"],
            ["username", "Instagram username (without @)", "YES"],
            ["status", "active, inactive, or blacklisted", "No (default: active)"],
            ["tier", "nano, micro, mid, macro, mega", "No"],
            ["categories", "Comma-separated (e.g. Fashion, Lifestyle)", "No"],
            ["tags", "Comma-separated (e.g. dubai, travel)", "No"],
            ["internal_notes", "Internal notes about the influencer", "No"],
            [],
            ["NOTE: Only username is required. Analytics (followers, engagement, etc.)"],
            ["populate automatically after import. Set pricing in the app after import."],
        ]
        for row_idx, row_data in enumerate(instructions, 1):
            for col_idx, val in enumerate(row_data, 1):
                cell = inst.cell(row=row_idx, column=col_idx, value=val)
                if row_idx == 1:
                    cell.font = Font(bold=True)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=influencer_import_template.xlsx"},
        )
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed. Run: pip install openpyxl")
    except Exception as e:
        logger.error(f"Error generating template: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# =============================================================================
# 13. EXCEL IMPORT
# =============================================================================

def _safe_str(v):
    return str(v).strip() if v is not None else None


def _safe_int(v):
    if v is None:
        return None
    try:
        return int(float(v))
    except (ValueError, TypeError):
        return None


def _safe_float(v):
    if v is None:
        return None
    try:
        return float(v)
    except (ValueError, TypeError):
        return None


def _dollars_to_cents(v):
    f = _safe_float(v)
    return int(round(f * 100)) if f is not None else None


def _parse_list(v):
    if not v:
        return []
    return [x.strip() for x in str(v).split(",") if x.strip()]


def _is_legacy_format(row_values: list) -> bool:
    """Detect legacy 30-column format: more than 10 non-None values means old template."""
    non_none = sum(1 for v in row_values if v is not None)
    return non_none > 10


def _parse_excel_row_legacy(row_values: list) -> dict:
    """Parse a row from the legacy 30-column Excel template."""
    username = _safe_str(row_values[0])
    if not username:
        return None

    return {
        "username": username.lstrip("@").lower(),
        "full_name": _safe_str(row_values[1]),
        "biography": _safe_str(row_values[2]),
        "is_verified": str(row_values[3]).upper() == "TRUE" if row_values[3] else False,
        "followers_count": _safe_int(row_values[4]),
        "following_count": _safe_int(row_values[5]),
        "posts_count": _safe_int(row_values[6]),
        "engagement_rate": _safe_float(row_values[7]),
        "avg_likes": _safe_int(row_values[8]),
        "avg_comments": _safe_int(row_values[9]),
        "avg_views": _safe_int(row_values[10]),
        "status": _safe_str(row_values[11]) or "active",
        "tier": _safe_str(row_values[12]),
        "categories": _parse_list(row_values[13]),
        "tags": _parse_list(row_values[14]),
        "internal_notes": _safe_str(row_values[15]),
        "cost_post_usd_cents": _dollars_to_cents(row_values[16]),
        "cost_story_usd_cents": _dollars_to_cents(row_values[17]),
        "cost_reel_usd_cents": _dollars_to_cents(row_values[18]),
        "cost_carousel_usd_cents": _dollars_to_cents(row_values[19]),
        "cost_video_usd_cents": _dollars_to_cents(row_values[20]),
        "cost_bundle_usd_cents": _dollars_to_cents(row_values[21]),
        "cost_monthly_usd_cents": _dollars_to_cents(row_values[22]),
        "sell_post_usd_cents": _dollars_to_cents(row_values[23]),
        "sell_story_usd_cents": _dollars_to_cents(row_values[24]),
        "sell_reel_usd_cents": _dollars_to_cents(row_values[25]),
        "sell_carousel_usd_cents": _dollars_to_cents(row_values[26]),
        "sell_video_usd_cents": _dollars_to_cents(row_values[27]),
        "sell_bundle_usd_cents": _dollars_to_cents(row_values[28]),
        "sell_monthly_usd_cents": _dollars_to_cents(row_values[29]),
    }


def _parse_excel_row(row_values: list) -> dict:
    """Parse a single Excel row, auto-detecting new 6-column or legacy 30-column format."""
    if _is_legacy_format(row_values):
        logger.info("Detected legacy 30-column Excel format")
        padded = list(row_values) + [None] * (len(LEGACY_TEMPLATE_COLUMNS) - len(row_values))
        return _parse_excel_row_legacy(padded)

    # New 6-column format: username, status, tier, categories, tags, internal_notes
    logger.info("Detected new 6-column Excel format")
    username = _safe_str(row_values[0])
    if not username:
        return None

    return {
        "username": username.lstrip("@").lower(),
        "status": _safe_str(row_values[1]) or "active",
        "tier": _safe_str(row_values[2]),
        "categories": _parse_list(row_values[3]),
        "tags": _parse_list(row_values[4]),
        "internal_notes": _safe_str(row_values[5]),
    }


@router.post("/influencer-database/import/excel")
async def import_from_excel(
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    current_user=Depends(require_admin()),
):
    """Import influencers from .xlsx file. Upserts by username, queues analytics for new imports."""
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="File must be .xlsx or .xls")

    try:
        from openpyxl import load_workbook
        from sqlalchemy import select

        contents = await file.read()

        if len(contents) > 5 * 1024 * 1024:
            raise HTTPException(status_code=400, detail="File too large (max 5 MB)")

        wb = load_workbook(io.BytesIO(contents), read_only=True, data_only=True)
        ws = wb.active

        rows = list(ws.iter_rows(min_row=2, values_only=True))  # skip header
        wb.close()

        if len(rows) > 500:
            raise HTTPException(status_code=400, detail="Maximum 500 rows per import")

        imported = 0
        updated = 0
        analytics_queued = 0
        analytics_skipped = 0
        errors = []
        imported_ids = []
        imported_usernames = []
        analytics_failures = []
        # Track new records that need analytics after commit
        new_records_for_analytics = []
        # Track all new usernames for ID lookup after commit
        all_new_usernames = []

        CHUNK_SIZE = 50
        from app.database.unified_models import InfluencerDatabase

        for chunk_start in range(0, len(rows), CHUNK_SIZE):
            chunk = rows[chunk_start:chunk_start + CHUNK_SIZE]
            try:
                for i, row in enumerate(chunk):
                    row_idx = chunk_start + i + 2  # +2 for 1-based + header
                    try:
                        padded = list(row) + [None] * max(0, len(LEGACY_TEMPLATE_COLUMNS) - len(row))
                        data = _parse_excel_row(padded)
                        if data is None:
                            continue

                        username = data.pop("username")

                        # Check if exists
                        result = await db.execute(
                            select(InfluencerDatabase).where(
                                InfluencerDatabase.username == username
                            )
                        )
                        existing = result.scalar_one_or_none()

                        if existing:
                            for key, val in data.items():
                                if val is not None:
                                    setattr(existing, key, val)
                            updated += 1
                            imported_usernames.append(username)
                        else:
                            # Check profiles table for existing analytics
                            profile_data = await InfluencerDatabaseService._get_from_profiles_table(db, username)
                            has_complete_analytics = bool(
                                profile_data and profile_data.get("ai_profile_analyzed_at")
                            )

                            # Merge profile data as defaults (Excel values take priority)
                            if profile_data:
                                profile_fields = {
                                    "full_name", "biography", "profile_image_url",
                                    "is_verified", "is_private", "followers_count",
                                    "following_count", "posts_count", "engagement_rate",
                                    "avg_likes", "avg_comments", "avg_views",
                                }
                                for field in profile_fields:
                                    if data.get(field) is None and profile_data.get(field) is not None:
                                        data[field] = profile_data[field]

                            a_status = "skipped" if has_complete_analytics else "pending"
                            inf = InfluencerDatabase(
                                username=username,
                                added_by=current_user.id,
                                analytics_status=a_status,
                                **{k: v for k, v in data.items() if v is not None},
                            )
                            db.add(inf)
                            imported += 1
                            all_new_usernames.append(username)
                            imported_usernames.append(username)

                            if has_complete_analytics:
                                analytics_skipped += 1
                            else:
                                new_records_for_analytics.append(username)

                    except Exception as row_err:
                        errors.append({"row": row_idx, "error": str(row_err)})

                await db.commit()
            except Exception as chunk_err:
                await db.rollback()
                chunk_end = min(chunk_start + CHUNK_SIZE, len(rows)) + 2
                logger.error(f"Chunk rows {chunk_start + 2}-{chunk_end} failed: {chunk_err}")
                errors.append({"row": f"{chunk_start + 2}-{chunk_end}", "error": f"Chunk failed: {chunk_err}"})

        # Look up IDs for ALL newly imported records
        for username in all_new_usernames:
            try:
                id_result = await db.execute(
                    text("SELECT id FROM influencer_database WHERE username = :u"),
                    {"u": username},
                )
                id_row = id_result.fetchone()
                if id_row:
                    imported_ids.append(str(id_row[0]))
            except Exception:
                pass

        # Queue analytics for new records that don't have complete analytics
        for username in new_records_for_analytics:
            try:
                id_result = await db.execute(
                    text("SELECT id FROM influencer_database WHERE username = :u"),
                    {"u": username},
                )
                id_row = id_result.fetchone()
                if id_row:
                    new_id = str(id_row[0])
                    await InfluencerDatabaseService.queue_analytics_for_influencer(
                        db, new_id, username, str(current_user.id)
                    )
                    analytics_queued += 1
            except Exception as q_err:
                logger.warning(f"Failed to queue analytics for @{username}: {q_err}")
                analytics_failures.append({"username": username, "reason": str(q_err)})

        # Get queue status
        queue_status = {"depth": 0, "max_depth": 200, "utilization_percent": 0.0}
        try:
            from app.core.job_queue import job_queue, QueueType
            depth = await job_queue._get_queue_depth(QueueType.API_QUEUE)
            max_depth = job_queue.queue_config[QueueType.API_QUEUE]['max_depth']
            queue_status = {
                "depth": depth,
                "max_depth": max_depth,
                "utilization_percent": round((depth / max_depth) * 100, 2) if max_depth else 0.0,
            }
        except Exception:
            pass

        return {
            "success": True,
            "data": {
                "imported": imported,
                "updated": updated,
                "analytics_queued": analytics_queued,
                "analytics_skipped": analytics_skipped,
                "errors": errors[:50],
                "total_processed": imported + updated + len(errors),
                "imported_ids": imported_ids,
                "imported_usernames": imported_usernames,
                "analytics_failures": analytics_failures,
                "queue_status": queue_status,
            },
            "message": f"Imported {imported}, updated {updated}, {analytics_queued} analytics queued, {len(errors)} errors",
        }

    except HTTPException:
        raise
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed. Run: pip install openpyxl")
    except Exception as e:
        await db.rollback()
        logger.error(f"Error importing Excel: {e}")
        raise HTTPException(status_code=500, detail=str(e))
